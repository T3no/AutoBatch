"""
This Program will automatically create the line items that will import into
the UPS worldship csv file

pip3 install openpyxl
pip3 install tkinter

"""
import tkinter as batch
import pandas as pd
import openpyxl
#from openpyxl.styles import Font


#open up my input excel file
wb=openpyxl.load_workbook('inputfile.xlsx')
#define worksheet 1
sheet = wb["Sheet1"]
#delete the uncessary columns from the input files.
#we only need the last 8 line items out of the 40 total
sheet.delete_cols(1,32)


    

#define and open the output work book
owb=openpyxl.load_workbook('outputfile.xlsx')
#define the output sheet
osheet = owb["Sheet1"]
#this will clear the worksheet by delete rows 1 to 40
osheet.delete_cols(1, 40)
       

def populate():
    osheet.delete_cols(1, 40)
    #this will fill the header page of the output file
    #this will be hardcoded inot the first row of any output file
    osheet.cell(row = 1, column = 1).value = 'RETURN'
    osheet.cell(row = 1, column = 2).value = 'RETURN_DECP'
    osheet.cell(row = 1, column = 3).value = 'RETURN_TYPE'
    osheet.cell(row = 1, column = 4).value = 'STORE_ID'
    osheet.cell(row = 1, column = 5).value = 'REF2'
    osheet.cell(row = 1, column = 6).value = 'Address2'
    osheet.cell(row = 1, column = 7).value = 'Company_To'
    osheet.cell(row = 1, column = 8).value = 'ATTENTION_TO'
    osheet.cell(row = 1, column = 9).value = 'ADDRESS'
    osheet.cell(row = 1, column = 10).value = 'CITY'
    osheet.cell(row = 1, column = 11).value = 'ST'
    osheet.cell(row = 1, column = 12).value = 'ZIP'
    osheet.cell(row = 1, column = 13).value = 'REF1'
    osheet.cell(row = 1, column = 14).value = 'PACKAGE_WEIGHT'
    osheet.cell(row = 1, column = 15).value = 'UPS_SERVICE'
    osheet.cell(row = 1, column = 16).value = 'BILLING'
    osheet.cell(row = 1, column = 17).value = 'PACKAGE_TYPE'
    osheet.cell(row = 1, column = 18).value = 'COMPANYFROM'
    osheet.cell(row = 1, column = 19).value = 'ATTENTION_FROM'
    osheet.cell(row = 1, column = 20).value = 'ADDRESS_FROM'
    osheet.cell(row = 1, column = 21).value = 'COUNTRY_FROM'
    osheet.cell(row = 1, column = 22).value = 'POSTAL_FROM'
    osheet.cell(row = 1, column = 23).value = 'CITY_FROM'
    osheet.cell(row = 1, column = 24).value = 'STATE_FROM'
    osheet.cell(row = 1, column = 25).value = 'PHONE_1'
    osheet.cell(row = 1, column = 26).value = 'UPS_ACC'
    osheet.cell(row = 1, column = 27).value = 'QVN'
    osheet.cell(row = 1, column = 28).value = 'QVN_EMAIL'
    
    cnt = 2
    hpcnt = 1
    apccnt = 1
    lcdcnt = 1
    ckitcnt = 1
    inp1cnt = 1
    inp2cnt =1
    inp3cnt = 1
    totcnt = 2
    s_method = shp_var.get()
    for i in range(sheet.max_row):
        if(hpValue.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,hpcnt)
            osheet['M' + str(totcnt)].value = 'HP'
            osheet['N' + str(totcnt)].value = 25
            cnt+=1
            hpcnt+=1
            totcnt+=1
        if(apcValue.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,apccnt)
            osheet['M' + str(totcnt)].value = 'APC'
            osheet['N' + str(totcnt)].value = 29
            cnt+=1
            apccnt+=1
            totcnt+=1
        if(lcdValue.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,lcdcnt)
            osheet['M' + str(totcnt)].value = 'LCD'
            osheet['N' + str(totcnt)].value = 16
            cnt+=1
            lcdcnt+=1
            totcnt+=1
        if(ckitValue.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,ckitcnt)
            osheet['M' + str(totcnt)].value = 'CKIT'
            osheet['N' + str(totcnt)].value = 25
            cnt+=1
            ckitcnt+=1
            totcnt+=1
        if(inp1Value.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,inp1cnt)
            osheet['M' + str(totcnt)].value = 'CKIT'
            osheet['N' + str(totcnt)].value = entryinp1.get()
            cnt+=1
            inp1cnt+=1
            totcnt+=1
        if(inp2Value.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,inp2cnt)
            osheet['M' + str(totcnt)].value = 'CKIT'
            osheet['N' + str(totcnt)].value = entryinp2.get()
            cnt+=1
            inp2cnt+=1
            totcnt+=1
        if(inp3Value.get() == 1):
            static_pop(sheet,osheet,cnt,s_method)
            add_pop(sheet,osheet,cnt,inp3cnt)
            osheet['M' + str(totcnt)].value = 'CKIT'
            osheet['N' + str(totcnt)].value = entryinp3.get()
            cnt+=1
            inp3cnt+=1
            totcnt+=1

                
    if(rtValue.get() == 1):
        hpcnt = 1
        apccnt = 1
        lcdcnt = 1
        ckitcnt = 1
        inp1cnt = 1
        inp2cnt =1
        inp3cnt = 1
        ship_desc = desc_var.get()
        return_depot = reloc.get() 
        for i in range(sheet.max_row):
            if(hpValue.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                        
                add_pop(sheet,osheet,cnt,hpcnt)
                osheet['M' + str(totcnt)].value = 'HP'
                osheet['N' + str(totcnt)].value = 25
                cnt+=1
                hpcnt+=1
                totcnt+=1
            if(apcValue.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                        
                add_pop(sheet,osheet,cnt,apccnt)
                osheet['M' + str(totcnt)].value = 'APC'
                osheet['N' + str(totcnt)].value = 29
                cnt+=1
                apccnt+=1
                totcnt+=1
            if(lcdValue.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                        
                add_pop(sheet,osheet,cnt,lcdcnt)
                osheet['M' + str(totcnt)].value = 'LCD'
                osheet['N' + str(totcnt)].value = 16
                cnt+=1
                lcdcnt+=1
                totcnt+=1
            if(ckitValue.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                    
                add_pop(sheet,osheet,cnt,ckitcnt)
                osheet['M' + str(totcnt)].value = 'CKIT'
                osheet['N' + str(totcnt)].value = 25
                cnt+=1
                ckitcnt+=1
                totcnt+=1
            if(inp1Value.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                    
                add_pop(sheet,osheet,cnt,inp1cnt)
                osheet['M' + str(totcnt)].value = 'CKIT'
                osheet['N' + str(totcnt)].value = entryinp1.get()
                cnt+=1
                inp1cnt+=1
                totcnt+=1
            if(inp2Value.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                add_pop(sheet,osheet,cnt,inp2cnt)
                osheet['M' + str(totcnt)].value = 'CKIT'
                osheet['N' + str(totcnt)].value = entryinp2.get()
                cnt+=1
                inp2cnt+=1
                totcnt+=1
            if(inp3Value.get() == 1):
                if (return_depot == 1):
                    itasca_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 2):
                    dellms_return_pop(sheet,osheet,cnt,ship_desc)
                elif (return_depot == 3):
                    memphis_return_pop(sheet,osheet,cnt,ship_desc)
                add_pop(sheet,osheet,cnt,inp3cnt)
                osheet['M' + str(totcnt)].value = 'CKIT'
                osheet['N' + str(totcnt)].value = entryinp3.get()
                cnt+=1
                inp3cnt+=1
                totcnt+=1

    owb.save('outputfile.xlsx')
    #after saving the outputfile.xlsx, it is read into the panda dataframe as string
    df = pd.read_excel("outputfile.xlsx","Sheet1",converters={"ZIP": str})
    #this will interate through column ZIP and look for zip codes that have
    #less than 5 charaters. if it finds a cell with a value less than 5
    #it will add 0 in front of it until the value is 5
    #this way the ups labels wont fail out for zip codes that start with 0
    for ind, row in df.iterrows():
        while(len(row['ZIP']) < 5):
              row['ZIP'] = '0' + row['ZIP']
              print(row['ZIP'])
        df.loc[ind,'ZIP'] = row['ZIP']
    df.to_csv('Camera_ProSys_CSV.csv',index=False)
        
    '''
    df_CSV = pd.read_csv('Camera_ProSys_CSV.csv',converters={"ZIP": str})
    for ind, row in df_CSV.iterrows():
        while(len(row['ZIP']) < 5):
              row['ZIP'] = '0' + row['ZIP']
              print(row['ZIP'])
        df_CSV.loc[ind,'ZIP'] = row['ZIP']
            
    df_CSV.to_csv('Camera_ProSys_CSV.csv',index=False)  
    '''

                
                  
        
#########################################################################################
#                                   GUI Design
#########################################################################################



#Height and Width variables
HEIGHT = 600
WIDTH = 800

#defining root of tinker
root = batch.Tk()

#my initials canvas sets the size of my window
canvas = batch.Canvas(root, height = HEIGHT, width = WIDTH, bg = '#EFF0F1')
canvas.pack()

#The program is split into left and right halves to make
#placing items a little easier
#the left frame has the line items such as apc,lcd,..ect
left_frame = batch.Frame(root, bd = 2 , bg = '#89CF89')
left_frame.place(relwidth = .6, relheight = 1)

#the right frame will have all the drop down menus
right_frame = batch.Frame(root, bd = 2 , bg = '#56AE56')
right_frame.place(relx = .6,relwidth = .4, relheight = 1)


# title label for the whole program 
title = batch.Label(left_frame, text = 'Auto Batch', font = 60,)
title.place(rely = 0, relx = .8)

# title label for line items
item = batch.Label(left_frame, text = 'Line Items', font = 40)
item.place(rely = .2, relx = .35)


#hp checkbox
#initalize variable for checkbox. This is our controller to add line items
#if the hp checkbox is checked when the save button is pressed hpValue will equal 1
hpValue = batch.IntVar()
#create check box and determine on value
hp = batch.Checkbutton(left_frame, text = 'HP', anchor = 'n', variable = hpValue, onvalue = 1)
#place check box on these x and y quardinates on the left frame 
hp.place(rely = .3, relx = .1)


#apc checkbox
#if the apc checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
apcValue = batch.IntVar()
apc = batch.Checkbutton(left_frame, text = 'APC', anchor = 'n', variable = apcValue, onvalue = 1)
apc.place(rely = .3, relx = .25)

#lcd checkbox
#if the lcd checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
lcdValue = batch.IntVar()
lcd = batch.Checkbutton(left_frame, text = 'LCD', anchor = 'n', variable = lcdValue, onvalue = 1)
lcd.place(rely = .3, relx = .4)

#c-kit checkbox
#if the c-kit checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
ckitValue = batch.IntVar()
ckit = batch.Checkbutton(left_frame, text = 'C-KIT', anchor = 'n', variable = ckitValue, onvalue = 1)
ckit.place(rely = .3, relx = .55)

#my return tracking check box
#if the return tracking checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
rtValue = batch.IntVar()
return_track = batch.Checkbutton(left_frame, text = 'Return Tracking',  variable = rtValue, onvalue = 1)
return_track.place(rely = .65, relx = .3)


#this will be the check box for items that are not predeterimend
#user will input the weight for entry
#if the input1 checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
inp1Value = batch.IntVar()
userinpt1 = batch.Checkbutton(left_frame, text = 'C-KIT', anchor = 'n',  variable = inp1Value, onvalue = 1)
userinpt1.place(rely = .4, relx = .15)

#this will be the check box 2 for items that are not predeterimend
#if the inp2 checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
inp2Value = batch.IntVar()
userinpt2 = batch.Checkbutton(left_frame, text = 'C-KIT', anchor = 'n',  variable = inp2Value, onvalue = 1)
userinpt2.place(rely = .4, relx = .3)

#this will be the check box 3 for items that are not predeterimend
#if the inp3 checkbox is checked when the save button is pressed hpValue will equal 1
#initalize variable for checkbox.
inp3Value = batch.IntVar()
userinpt3 = batch.Checkbutton(left_frame, text = 'C-KIT', anchor = 'n',  variable = inp3Value, onvalue = 1)
userinpt3.place(rely = .4, relx = .45)

#this is the weight entry box for user input 1
entryinp1 = batch.Entry(left_frame )
entryinp1.place(rely = .47, relx = .15, relwidth = 0.115)

#this is the weight entry box for user input 2
entryinp2 = batch.Entry(left_frame)
entryinp2.place(rely = .47, relx = .3, relwidth = 0.115)

#this is the weight entry box for user input 3
entryinp3 = batch.Entry(left_frame)
entryinp3.place(rely = .47, relx = .45, relwidth = 0.115)

#this is the execute/quit button
#because i want this button to save the output file, i passed in the output excel workbook
button = batch.Button(left_frame, text = "Exit", font=40, command = lambda: execute(owb))
button.place(relx=0.5, rely = 0.8,relwidth = .3, relheight = .1)

#this is the save button
#this button will call the populate funtion
#when this button is pressed we will gather info from all the check boxes and drop downs
save = batch.Button(left_frame, text = "Save", font=40, command = populate)
save.place(relx=0.1, rely = 0.8,relwidth = .3, relheight = .1)


#my return location drop down box
reloc = batch.IntVar()
reloc.set(1)
loc_lab = batch.Label(right_frame, text = "Choose Return Location")
loc_lab.place(rely = .45, relx = .3)
loc1 = batch.Radiobutton(right_frame, text="Itasca", variable=reloc, value=1)
loc1.place(rely = .50, relx = .4)
loc2 = batch.Radiobutton(right_frame, text="Dell MS", variable=reloc, value=2)
loc2.place( rely = .545, relx = .4 )
loc3 = batch.Radiobutton(right_frame, text="Memphis", variable=reloc, value=3)
loc3.place( rely = .59, relx = .4)
   
   


#radio button for return description

desc_var = batch.IntVar()
desc_var.set(1)
desc_lab = batch.Label(right_frame, text = "Return Description")
desc_lab.place(rely = .7, relx = .3)
desc1 = batch.Radiobutton(right_frame, text="HP Swap", variable=desc_var, value=1)
desc1.place(rely = .75, relx = .4)
desc2 = batch.Radiobutton(right_frame, text="SC Digital Conversion", variable=desc_var, value=2)
desc2.place( rely = .795, relx = .4 )
desc3 = batch.Radiobutton(right_frame, text="Lease Roll", variable=desc_var, value=3)
desc3.place( rely = .84, relx = .4)
desc3 = batch.Radiobutton(right_frame, text="Hot Swap Return", variable=desc_var, value=4)
desc3.place( rely = .885, relx = .4)
desc4 = batch.Radiobutton(right_frame, text="HP Swap/30 Day Hold", variable=desc_var, value=5)
desc4.place( rely = .93, relx = .4)



    
#radio button for shipping method
shp_var = batch.IntVar()
shp_var.set(1)
shp_lab = batch.Label(right_frame, text = "Choose A Shipping Method")
shp_lab.place(rely = .11, relx = .3)
R1 = batch.Radiobutton(right_frame, text="Ground", variable=shp_var, value=1)
R1.place(rely = .17, relx = .4)
R2 = batch.Radiobutton(right_frame, text="Overnight", variable=shp_var, value=2)
R2.place( rely = .215, relx = .4 )
R3 = batch.Radiobutton(right_frame, text="2 Day Air", variable=shp_var, value=3)
R3.place( rely = .26, relx = .4)
R3 = batch.Radiobutton(right_frame, text="3 Day Select", variable=shp_var, value=4)
R3.place( rely = .305, relx = .4)

    

#############################################################################################################################
#                                                Funtions out of main
#############################################################################################################################



#This funtion auto fills the cells that have static values
#This function auto populates through a counter
#Every time we make a line item through the gui/populate() it will populate a line item
#we pass the cnt couner variable from the populate funtion on the top
def static_pop(sheet,osheet,cnt,s_method):
    osheet['C' + str(cnt)].value = 'PRL'
    osheet['D' + str(cnt)].value = 'STORE ID'
    osheet['P' + str(cnt)].value = 'shp'
    osheet['Q' + str(cnt)].value = 'package'
    osheet['R' + str(cnt)].value = 'FAHAD SYED/SA7384'
    osheet['S' + str(cnt)].value = 'AT&T'
    osheet['T' + str(cnt)].value = '1100 MAPLEWOOD DR'
    osheet['U' + str(cnt)].value = 'UNITED STATES'
    osheet['V' + str(cnt)].value = '60143'
    osheet['W' + str(cnt)].value = 'ITASCA'
    osheet['X' + str(cnt)].value = 'IL'
    osheet['Y' + str(cnt)].value = '6309314131'
    osheet['Z' + str(cnt)].value = '7F4049'
    osheet['AA' + str(cnt)].value = 'Y'
    osheet['AB' + str(cnt)].value = 'SA7384@ATT.COM'
    if s_method == 1:
        osheet['O' + str(cnt)].value = 'GND'
    elif s_method == 2:
        osheet['O' + str(cnt)].value = '1DA'
    elif s_method == 3:
        osheet['O' + str(cnt)].value = '2DA'
    elif s_method == 4:
        osheet['O' + str(cnt)].value = '3DS'


#This funtion auto fills the cells that have static values if a return label is required for the Itasca Depot
#This function auto populates through a counter
#Every time we make a line item through the gui/populate() it will populate a line item
#we pass the cnt couner variable from the populate funtion on the top
def itasca_return_pop(sheet,osheet,cnt,ship_desc):
    osheet['A' + str(cnt)].value = 'Y'
    osheet['C' + str(cnt)].value = 'PRL'
    osheet['D' + str(cnt)].value = 'STORE ID'
    osheet['P' + str(cnt)].value = 'shp'
    osheet['Q' + str(cnt)].value = 'package'
    osheet['R' + str(cnt)].value = 'FAHAD SYED/SA7384'
    osheet['S' + str(cnt)].value = 'AT&T'
    osheet['T' + str(cnt)].value = '1100 MAPLEWOOD DR'
    osheet['U' + str(cnt)].value = 'UNITED STATES'
    osheet['V' + str(cnt)].value = '60143'
    osheet['W' + str(cnt)].value = 'ITASCA'
    osheet['X' + str(cnt)].value = 'IL'
    osheet['Y' + str(cnt)].value = '6309314131'
    osheet['Z' + str(cnt)].value = '7F4049'
    osheet['O' + str(cnt)].value = 'GND'
    if ship_desc == 1:
        osheet['B' + str(cnt)].value = 'HP Swap'
    elif ship_desc == 2:
        osheet['B' + str(cnt)].value = 'SC Digital Conversion'
    elif ship_desc == 3:
        osheet['B' + str(cnt)].value = 'Lease Roll'
    elif ship_desc == 4:
        osheet['B' + str(cnt)].value = 'Hot Swap Return'
    elif ship_desc == 5:
        osheet['B' + str(cnt)].value = 'HP Swap/30 Day Hold'


#This funtion auto fills the cells that have static values if a return label is required for the DelLMS Depot
#This function auto populates through a counter
#Every time we make a line item through the gui/populate() it will populate a line item
#we pass the cnt couner variable from the populate funtion on the top
def dellms_return_pop(sheet,osheet,cnt,ship_desc):
    osheet['A' + str(cnt)].value = 'Y'
    osheet['C' + str(cnt)].value = 'PRL'
    osheet['D' + str(cnt)].value = 'STORE ID'
    osheet['P' + str(cnt)].value = 'shp'
    osheet['Q' + str(cnt)].value = 'package'
    osheet['R' + str(cnt)].value = 'ATT COR'
    osheet['S' + str(cnt)].value = 'Dell MS Parts'
    osheet['T' + str(cnt)].value = '5985 Cabot Parkway'
    osheet['U' + str(cnt)].value = 'UNITED STATES'
    osheet['V' + str(cnt)].value = '30005'
    osheet['W' + str(cnt)].value = 'ALPHARETTA'
    osheet['X' + str(cnt)].value = 'GA'
    osheet['Y' + str(cnt)].value = '6782681310'
    osheet['Z' + str(cnt)].value = '7F4049'
    osheet['O' + str(cnt)].value = 'GND'
    if ship_desc == 1:
        osheet['B' + str(cnt)].value = 'HP Swap'
    elif ship_desc == 2:
        osheet['B' + str(cnt)].value = 'SC Digital Conversion'
    elif ship_desc == 3:
        osheet['B' + str(cnt)].value = 'Lease Roll'
    elif ship_desc == 4:
        osheet['B' + str(cnt)].value = 'Hot Swap Return'
    elif ship_desc == 5:
        osheet['B' + str(cnt)].value = 'HP Swap/30 Day Hold'

#This funtion auto fills the cells that have static values if a return label is required for the Memphis Depot
#This function auto populates through a counter
#Every time we make a line item through the gui/populate() it will populate a line item
#we pass the cnt couner variable from the populate funtion on the top
def memphis_return_pop(sheet,osheet,cnt,ship_desc):
    osheet['A' + str(cnt)].value = 'Y'
    osheet['C' + str(cnt)].value = 'PRL'
    osheet['D' + str(cnt)].value = 'STORE ID'
    osheet['P' + str(cnt)].value = 'shp'
    osheet['Q' + str(cnt)].value = 'package'
    osheet['R' + str(cnt)].value = 'DD80004/ATT'
    osheet['S' + str(cnt)].value = 'DEPOT'
    osheet['T' + str(cnt)].value = '6400 SHELBY VIEW DR'
    osheet['U' + str(cnt)].value = 'UNITED STATES'
    osheet['V' + str(cnt)].value = '38134'
    osheet['W' + str(cnt)].value = 'BARTLETT'
    osheet['X' + str(cnt)].value = 'TN'
    osheet['Y' + str(cnt)].value = '8472483334'
    osheet['Z' + str(cnt)].value = '7F4049'
    osheet['O' + str(cnt)].value = 'GND'
    if ship_desc == 1:
        osheet['B' + str(cnt)].value = 'HP Swap'
    elif ship_desc == 2:
        osheet['B' + str(cnt)].value = 'SC Digital Conversion'
    elif ship_desc == 3:
        osheet['B' + str(cnt)].value = 'Lease Roll'
    elif ship_desc == 4:
        osheet['B' + str(cnt)].value = 'Hot Swap Return'
    elif ship_desc == 5:
        osheet['B' + str(cnt)].value = 'HP Swap/30 Day Hold'




#This funtions  fills specific cells in the output file with cells from the input file
#Every time we make a line item through the gui/populate() it will populate a line item
#we pass the cnt and popcnt counter variable from the populate funtion on the top
#.strip() will remove any trailing/leading spaces we copy in
def add_pop(sheet,osheet,cnt,popcnt):
    osheet['E' + str(cnt)].value = str(sheet['A'+ str(popcnt)].value).strip()
    osheet['F' + str(cnt)].value = str(sheet['B'+ str(popcnt)].value).strip()
    osheet['G' + str(cnt)].value = str(sheet['C'+ str(popcnt)].value).strip()
    osheet['H' + str(cnt)].value = str(sheet['D'+ str(popcnt)].value).strip()
    osheet['I' + str(cnt)].value = str(sheet['E'+ str(popcnt)].value).strip()
    osheet['J' + str(cnt)].value = str(sheet['F'+ str(popcnt)].value).strip()
    osheet['K' + str(cnt)].value = str(sheet['G'+ str(popcnt)].value).strip()
    osheet['L' + str(cnt)].value = str(sheet['H'+ str(popcnt)].value).strip()



#This funtion excutes as soon as we press the execute button
#It is responsible for saving the changes we made to the output file
#It also quits the program
def execute(owb):
    #owb.save('outputfile.xlsx')
    root.destroy()

    
root.mainloop()
